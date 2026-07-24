#!/usr/bin/env python3
"""把订单 xlsx 的每一行整理为独立文件夹，并下载相关素材。"""

from __future__ import annotations

import argparse
import html
import mimetypes
import os
import re
import shutil
import ssl
import sys
import tarfile
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen

import openpyxl


INVALID_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
# 以 http(s):// 起点定位 URL；分隔符不枚举，避免漏掉业务侧新分隔写法。
URL_START_RE = re.compile(r"https?://", re.IGNORECASE)
# 合法下载 URL 末尾通常落在字母数字或常见路径/查询字符上。
URL_TAIL_OK_RE = re.compile(r"[A-Za-z0-9/=%~_-]$")
# 兼容旧调用：页面效果图探测等仍按“找第一个 URL”使用。
URL_RE = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)
ARCHIVE_SUFFIXES = (".tar.gz", ".tar.bz2", ".tgz", ".tbz2", ".zip", ".tar", ".7z")
FAILURE_LOG_NAME = "失败日志.txt"
LINK_FILE_NAME = "link.txt"
# macOS 上微信/系统沙盒目录通常不可由其他程序写入。
RESTRICTED_OUTPUT_MARKERS = (
    "/Library/Containers/",
    "/Library/Group Containers/",
)
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
HTTP_TIMEOUT = 90
BROWSER_TIMEOUT = 45


def ssl_context() -> ssl.SSLContext:
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except Exception:
        return ssl.create_default_context()


SSL_CONTEXT = ssl_context()


@dataclass
class ProgressEvent:
    order_index: int
    order_total: int
    order_no: str
    item_index: int = 0
    item_total: int = 0
    message: str = ""


@dataclass
class DownloadFailure:
    """单个素材下载/解压失败的详细记录，用于写入订单目录内失败日志。"""

    order_no: str
    category: str
    url: str
    reason: str
    browser_used: bool = False
    direct_error: str = ""
    stage: str = "download"  # download | extract
    file_name: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


ProgressCallback = Callable[[ProgressEvent], None]


def emit(callback: ProgressCallback | None, event: ProgressEvent) -> None:
    if callback:
        callback(event)


def path_is_restricted(path: Path) -> bool:
    """判断路径是否位于常见不可写沙盒目录（如微信文件目录）。"""
    try:
        text = str(path.expanduser().resolve())
    except OSError:
        text = str(path.expanduser())
    normalized = text.replace("\\", "/")
    return any(marker in normalized for marker in RESTRICTED_OUTPUT_MARKERS)


def probe_directory_writable(path: Path) -> None:
    """确认目录可创建且可写入；失败时抛出带操作建议的错误。"""
    path = path.expanduser()
    try:
        path.mkdir(parents=True, exist_ok=True)
        probe = path / f".xlsx2files_write_test_{os.getpid()}"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
    except OSError as exc:
        raise RuntimeError(
            f"无法写入输出目录：{path}\n"
            f"原因：{exc}\n"
            "请改选桌面、下载文件夹或其他有写权限的位置。"
            "从微信等聊天软件目录打开的文件，默认输出路径可能受系统保护。"
        ) from exc


def suggest_output_dir(xlsx_path: Path) -> Path:
    """为 Excel 选择合适的默认输出目录；避开微信等受保护路径。"""
    xlsx_path = xlsx_path.expanduser()
    name = f"{xlsx_path.stem}_输出"
    beside = xlsx_path.with_name(name)
    if not path_is_restricted(beside):
        parent = beside.parent
        if parent.is_dir():
            try:
                probe_directory_writable(parent)
                return beside
            except RuntimeError:
                pass

    for base in (Path.home() / "Desktop", Path.home() / "Downloads", Path.home() / "Documents"):
        if base.is_dir() and not path_is_restricted(base):
            candidate = base / name
            try:
                probe_directory_writable(base)
                return candidate
            except RuntimeError:
                continue
    return Path.cwd() / name


def safe_name(value: object, fallback: str) -> str:
    name = INVALID_FILENAME.sub("_", str(value or "").strip()).rstrip(". ")
    return name or fallback


def download_category(header: str, column: int) -> str:
    """将“效果图下载地址”转换为稳定的分类目录名“效果图”。"""
    name = re.sub(r"下载地址$", "", header).strip()
    return safe_name(name, f"下载内容_{column}")


def flatten_legacy_download_dirs(folder: Path, directory_names: set[str]) -> int:
    """迁移旧版本遗留的分类目录，将其顶层内容移动到订单根目录。"""
    moved = 0
    for name in directory_names:
        legacy = folder / name
        if not legacy.is_dir():
            continue
        for source in list(legacy.iterdir()):
            destination = unique_path(folder / safe_name(source.name, "下载内容"))
            shutil.move(str(source), destination)
            moved += 1
        legacy.rmdir()
    return moved


def prepare_order_folder(output_root: Path, folder_name: str) -> Path:
    """复用上次的失败目录，重跑时先恢复正常名称以便重新验证。"""
    normal = output_root / folder_name
    failed = output_root / f"{folder_name}_下载失败"
    if not normal.exists() and failed.is_dir():
        failed.rename(normal)
    normal.mkdir(parents=True, exist_ok=True)
    return normal


def format_failure_log(folder_name: str, failures: list[DownloadFailure]) -> str:
    """生成订单目录内可读的详细失败日志。"""
    lines = [
        f"订单号: {folder_name}",
        f"失败项数: {len(failures)}",
        f"记录时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "说明: link.txt 仅含失败 URL，便于人工复制下载；本文件记录每次失败的详细原因。",
        "",
    ]
    stage_labels = {"download": "下载", "extract": "解压"}
    for index, item in enumerate(failures, start=1):
        lines.extend(
            [
                f"-------- 失败项 {index} --------",
                f"时间: {item.timestamp}",
                f"阶段: {stage_labels.get(item.stage, item.stage)}",
                f"类别: {item.category}",
            ]
        )
        if item.url:
            lines.append(f"URL: {item.url}")
        if item.file_name:
            lines.append(f"文件: {item.file_name}")
        lines.append(f"原因: {item.reason}")
        lines.append(f"已启动浏览器回退: {'是' if item.browser_used else '否'}")
        if item.direct_error:
            lines.append(f"直连失败原因: {item.direct_error}")
        lines.append("")
    return "\n".join(lines)


def finalize_order_folder(folder: Path, folder_name: str, failures: list[DownloadFailure]) -> Path:
    """下载失败时写 link.txt / 失败日志 并给目录加后缀；成功时清除旧标记。"""
    link_file = folder / LINK_FILE_NAME
    log_file = folder / FAILURE_LOG_NAME
    download_failures = [item for item in failures if item.stage == "download" and item.url]
    if download_failures:
        unique_urls = list(dict.fromkeys(item.url for item in download_failures))
        link_file.write_text("\n".join(unique_urls) + "\n", encoding="utf-8-sig")
        log_file.write_text(format_failure_log(folder_name, failures), encoding="utf-8-sig")
        target = folder.with_name(f"{folder_name}_下载失败")
        if folder != target:
            if target.exists():
                target = unique_path(target)
            folder.rename(target)
        return target
    if link_file.exists():
        link_file.unlink()
    if log_file.exists():
        log_file.unlink()
    return folder


def split_filename(name: str) -> tuple[str, str]:
    lower = name.lower()
    for suffix in ARCHIVE_SUFFIXES:
        if lower.endswith(suffix):
            return name[: -len(suffix)], name[-len(suffix) :]
    path = Path(name)
    return path.stem, path.suffix


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = split_filename(path.name)
    index = 2
    while True:
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def split_url_candidates(text: str) -> list[str]:
    """从任意拼接文本中切出候选 URL。

    不依赖固定分隔符列表：凡是下一个 ``http://`` / ``https://`` 出现处即切开，
    因此 ``|``、``;``、中文标点、换行、甚至未知符号夹在两条绝对地址之间都能拆开；
    单个 URL 内部的 ``,``、``&`` 等合法字符会保留。
    """
    starts = [match.start() for match in URL_START_RE.finditer(text)]
    if not starts:
        return []

    candidates: list[str] = []
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else len(text)
        chunk = text[start:end]
        # 空白几乎不会出现在未编码 URL 中，作为硬边界截断更安全。
        chunk = re.split(r"\s+", chunk, maxsplit=1)[0]
        # 去掉粘在两条 URL 之间的任意分隔残留（|,;,@,@,中文标点等）。
        while chunk and not URL_TAIL_OK_RE.search(chunk):
            chunk = chunk[:-1]
        if chunk and URL_START_RE.match(chunk):
            candidates.append(chunk)
    return candidates


def extract_urls(value: object) -> list[str]:
    """提取下载 URL，并展开 Printerval 的 design_urls 参数。"""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []

    found: list[str] = []
    for url in split_url_candidates(text):
        parsed = urlparse(url)
        query = parse_qs(parsed.query)
        design_urls = query.get("design_urls", [])
        if design_urls:
            for item in ",".join(design_urls).split(","):
                item = unquote(item.strip())
                if item:
                    # Printerval 的相对素材路径实际位于 assets 子域名。
                    asset_host = (
                        "assets.printerval.com"
                        if parsed.netloc.endswith("printerval.com")
                        else parsed.netloc
                    )
                    found.append(item if item.startswith("http") else f"{parsed.scheme}://{asset_host}{item}")
        else:
            found.append(url)
    return list(dict.fromkeys(found))


def discover_page_images(page_url: str) -> list[str]:
    """发现页面展示但没有下载按钮的原始产品效果图。"""
    parsed = urlparse(page_url)
    if not (parsed.netloc.endswith("printerval.com") and "/folder-design" in parsed.path):
        return []
    request = Request(page_url, headers={"User-Agent": USER_AGENT})
    try:
        with urlopen(request, timeout=60, context=SSL_CONTEXT) as response:
            page = html.unescape(response.read().decode("utf-8", errors="ignore")).replace("\\/", "/")
    except Exception as exc:
        print(f"    警告：无法解析页面效果图：{exc}", file=sys.stderr)
        return []

    candidates = re.findall(
        r'https://cdn\.printerval\.com/image/[^\s"\'<>]+?\.(?:png|jpe?g|webp)(?:\?[^\s"\'<>]*)?',
        page,
        flags=re.I,
    )
    result = []
    for candidate in candidates:
        # /image/960x960/... 是同一图片的缩略图，保留无尺寸段的原图。
        if re.search(r"/image/\d+x\d+/", candidate, re.I):
            continue
        if candidate not in result:
            result.append(candidate)
    return result[:1]


def filename_from_response(url: str, headers: object, index: int) -> str:
    disposition = headers.get("Content-Disposition", "")
    match = re.search(r"filename\*=UTF-8''([^;]+)|filename=\"?([^\";]+)", disposition, re.I)
    if match:
        return safe_name(unquote(match.group(1) or match.group(2)), f"download_{index}")
    name = Path(unquote(urlparse(url).path)).name
    if name:
        return safe_name(name, f"download_{index}")
    content_type = headers.get_content_type() if hasattr(headers, "get_content_type") else ""
    return f"download_{index}{mimetypes.guess_extension(content_type) or ''}"


def download_http(url: str, target_dir: Path, index: int) -> list[Path]:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=HTTP_TIMEOUT, context=SSL_CONTEXT) as response:
        final_url = response.geturl()
        filename = filename_from_response(final_url, response.headers, index)
        target = unique_path(target_dir / filename)
        with target.open("wb") as output:
            shutil.copyfileobj(response, output)
    return [target]


def download_google_drive(url: str, target_dir: Path) -> list[Path]:
    try:
        import gdown
    except ImportError as exc:
        raise RuntimeError("Google Drive 下载需要安装 gdown（pip install -r requirements.txt）") from exc

    # 先下载到临时目录，再平铺移动到订单目录；这样不会留下 Google Drive
    # 文件夹层级，也能统一处理同名文件冲突。
    with tempfile.TemporaryDirectory(prefix=".gdrive-", dir=target_dir) as staging_value:
        staging = Path(staging_value)
        if "/folders/" in url:
            gdown.download_folder(url=url, output=str(staging), quiet=False, use_cookies=False)
        else:
            parsed = urlparse(url)
            file_id = parse_qs(parsed.query).get("id", [None])[0]
            if not file_id:
                match = re.search(r"/d/([^/]+)", parsed.path)
                file_id = match.group(1) if match else None
            if not file_id:
                raise RuntimeError(f"无法识别 Google Drive 文件 ID：{url}")
            try:
                output = gdown.download(id=file_id, output=str(staging) + os.sep, quiet=False)
            except Exception as file_error:
                # open?id=... 既可能指向文件，也可能重定向到文件夹。
                folder_url = f"https://drive.google.com/drive/folders/{file_id}"
                print("    Google Drive 文件模式失败，按文件夹重试")
                try:
                    gdown.download_folder(url=folder_url, output=str(staging), quiet=False, use_cookies=False)
                    output = True
                except Exception as folder_error:
                    raise RuntimeError(
                        f"Google Drive 文件和文件夹模式均失败；文件错误：{file_error}；文件夹错误：{folder_error}"
                    ) from folder_error
            if not output:
                raise RuntimeError("Google Drive 未返回下载文件")

        moved: list[Path] = []
        for source in sorted(p for p in staging.rglob("*") if p.is_file()):
            destination = unique_path(target_dir / safe_name(source.name, "google_drive_file"))
            shutil.move(str(source), destination)
            moved.append(destination)
        return moved


def download_url(url: str, target_dir: Path, index: int) -> list[Path]:
    if urlparse(url).netloc.lower().endswith("drive.google.com"):
        return download_google_drive(url, target_dir)
    return download_http(url, target_dir, index)


def download_with_browser(
    url: str,
    target_dir: Path,
    profile_dir: Path,
    headless: bool,
    timeout_seconds: int = BROWSER_TIMEOUT,
) -> list[Path]:
    """使用本机 Chrome/Edge 会话触发下载，作为直连失败后的回退。"""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        bundled_browsers = Path(sys._MEIPASS) / ".playwright-browsers"
    else:
        bundled_browsers = Path(__file__).resolve().parent / ".playwright-browsers"
    if bundled_browsers.exists():
        os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(bundled_browsers))

    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("浏览器回退需要安装 playwright（pip install -r requirements.txt）") from exc

    profile_dir.mkdir(parents=True, exist_ok=True)
    timeout_ms = timeout_seconds * 1000
    before = {p.resolve() for p in target_dir.rglob("*") if p.is_file()}
    with sync_playwright() as playwright:
        context = None
        errors: list[str] = []
        # 无头模式优先使用 exe 内置 Chromium；可见模式优先使用系统 Chrome/Edge。
        channels = (None, "chrome", "msedge") if headless else ("chrome", "msedge", None)
        for channel in channels:
            try:
                context = playwright.chromium.launch_persistent_context(
                    user_data_dir=str(profile_dir),
                    channel=channel,
                    headless=headless,
                    accept_downloads=True,
                    downloads_path=str(target_dir),
                    args=["--disable-blink-features=AutomationControlled"],
                    timeout=timeout_ms,
                )
                break
            except Exception as exc:
                errors.append(f"{channel or 'chromium'}: {exc}")
        if context is None:
            raise RuntimeError("无法启动 Chrome/Edge：" + " | ".join(errors))

        try:
            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(timeout_ms)
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)

            # Google Drive 预览页可能使用中文或英文按钮。
            selectors = [
                '[aria-label="下载"]',
                '[aria-label^="下载"]',
                '[aria-label="Download"]',
                '[aria-label^="Download"]',
                "a[download]",
            ]
            for selector in selectors:
                buttons = page.locator(selector)
                if buttons.count() == 0:
                    continue
                try:
                    with page.expect_download(timeout=timeout_ms) as event:
                        buttons.first.click()
                    download = event.value
                    destination = unique_path(target_dir / safe_name(download.suggested_filename, "browser_download"))
                    download.save_as(str(destination))
                    return [destination]
                except PlaywrightTimeoutError:
                    continue

            # 普通资源页没有下载按钮时，用已建立的浏览器会话重新请求原地址。
            response = context.request.get(url, timeout=timeout_ms)
            if response.ok:
                headers = response.headers
                disposition = headers.get("content-disposition", "")
                match = re.search(r"filename\*=UTF-8''([^;]+)|filename=\"?([^\";]+)", disposition, re.I)
                name = unquote((match.group(1) or match.group(2)) if match else Path(urlparse(url).path).name)
                destination = unique_path(target_dir / safe_name(name, "browser_download"))
                destination.write_bytes(response.body())
                return [destination]
            raise RuntimeError(f"浏览器已打开页面，但未找到可下载按钮（HTTP {response.status}）")
        finally:
            context.close()

    return [p for p in target_dir.rglob("*") if p.is_file() and p.resolve() not in before]


def is_archive_file(path: Path) -> bool:
    lower = path.name.lower()
    if lower.endswith(ARCHIVE_SUFFIXES):
        return True
    try:
        if zipfile.is_zipfile(path):
            return True
        if tarfile.is_tarfile(path):
            return True
    except Exception:
        return False
    return False


def safe_extract_archive(archive: Path) -> Path | None:
    if not is_archive_file(archive):
        return None
    lower = archive.name.lower()
    stem = archive.name
    for suffix in ARCHIVE_SUFFIXES:
        if lower.endswith(suffix):
            stem = archive.name[: -len(suffix)]
            break
    else:
        stem = archive.stem

    destination = unique_path(archive.parent / safe_name(stem, "archive"))
    destination.mkdir(parents=True)
    root = destination.resolve()

    def validate(names: list[str]) -> None:
        for name in names:
            resolved = (destination / name).resolve()
            if root != resolved and root not in resolved.parents:
                raise RuntimeError(f"压缩包包含不安全路径：{name}")

    try:
        if zipfile.is_zipfile(archive) or lower.endswith(".zip"):
            with zipfile.ZipFile(archive) as package:
                validate(package.namelist())
                package.extractall(destination)
        elif tarfile.is_tarfile(archive) or lower.endswith((".tar", ".tar.gz", ".tgz", ".tar.bz2", ".tbz2")):
            with tarfile.open(archive) as package:
                validate(package.getnames())
                package.extractall(destination, filter="data")
        elif lower.endswith(".7z"):
            try:
                import py7zr
            except ImportError as exc:
                raise RuntimeError("解压 7z 需要安装 py7zr（pip install -r requirements.txt）") from exc
            with py7zr.SevenZipFile(archive) as package:
                validate(package.getnames())
                package.extractall(destination)
        else:
            shutil.rmtree(destination)
            return None
    except Exception:
        if destination.exists():
            shutil.rmtree(destination, ignore_errors=True)
        raise
    return destination


def image_map(sheet: object, qr_column: int) -> dict[int, object]:
    result: dict[int, object] = {}
    for image in sheet._images:
        anchor = getattr(image.anchor, "_from", None)
        if anchor and anchor.col + 1 == qr_column:
            result.setdefault(anchor.row + 1, image)
    return result


def write_qr(image: object, folder: Path) -> Path:
    extension = (getattr(image, "format", None) or "png").lower()
    path = folder / f"二维码.{extension}"
    path.write_bytes(image._data())
    return path


def write_text_summary(sheet: object, row: int, headers: list[str], excluded: set[int], folder: Path) -> Path:
    lines = []
    for column, header in enumerate(headers, start=1):
        value = sheet.cell(row, column).value
        if column not in excluded and value not in (None, ""):
            lines.append(f"{header}: {value}")
    path = folder / "信息.txt"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return path


def collect_cell_urls(cell_value: object) -> list[str]:
    """提取单元格全部下载链接，并补充 Printerval 页面效果图。"""
    urls = extract_urls(cell_value)
    if not cell_value:
        return urls
    page_url_match = URL_RE.search(str(cell_value))
    if not page_url_match:
        return urls
    extras = discover_page_images(page_url_match.group(0))
    if extras:
        print(f"  发现无下载按钮的页面效果图：{len(extras)} 张")
        urls = list(dict.fromkeys(extras + urls))
    return urls


def process(args: argparse.Namespace, progress_callback: ProgressCallback | None = None) -> int:
    workbook = openpyxl.load_workbook(args.xlsx, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    headers = [str(cell.value or f"未命名列{cell.column}").strip() for cell in sheet[1]]
    qr_columns = [i for i, name in enumerate(headers, 1) if name == "二维码"]
    download_columns = [i for i, name in enumerate(headers, 1) if "下载地址" in name]
    if not qr_columns:
        print("警告：找不到名为“二维码”的列，将跳过二维码提取", file=sys.stderr)
    if not download_columns and not args.skip_downloads:
        raise RuntimeError("找不到名称含“下载地址”的列")

    output_root = Path(args.output)
    probe_directory_writable(output_root)
    qr_images = image_map(sheet, qr_columns[0]) if qr_columns else {}
    last_row = min(sheet.max_row, 1 + args.limit) if args.limit else sheet.max_row
    order_total = max(0, last_row - 1)
    failures = 0
    cancel_event = getattr(args, "cancel_event", None)

    for row in range(2, last_row + 1):
        if cancel_event and cancel_event.is_set():
            print("已请求停止，任务结束")
            return 130
        order_index = row - 1
        folder_name = safe_name(sheet.cell(row, 1).value, f"第{row}行")
        folder = prepare_order_folder(output_root, folder_name)
        print(f"[{order_index}/{order_total}] {folder_name}")
        emit(
            progress_callback,
            ProgressEvent(
                order_index=order_index,
                order_total=order_total,
                order_no=folder_name,
                message=f"开始处理 {folder_name}",
            ),
        )
        legacy_names = {download_category(headers[column - 1], column) for column in download_columns}
        migrated = flatten_legacy_download_dirs(folder, legacy_names)
        if migrated:
            print(f"  已将旧分类目录中的 {migrated} 项内容平铺到订单目录")
        write_text_summary(sheet, row, headers, set(qr_columns + download_columns), folder)
        if qr_columns:
            if row in qr_images:
                write_qr(qr_images[row], folder)
            else:
                print(f"  警告：第 {row} 行没有找到二维码图片", file=sys.stderr)
                failures += 1

        download_index = 1
        failure_records: list[DownloadFailure] = []
        pending_urls: list[tuple[str, str]] = []
        for column in ([] if args.skip_downloads else download_columns):
            category = headers[column - 1]
            urls = collect_cell_urls(sheet.cell(row, column).value)
            for url in urls:
                pending_urls.append((category, url))

        item_total = len(pending_urls)
        for item_index, (category, url) in enumerate(pending_urls, start=1):
            if cancel_event and cancel_event.is_set():
                print("已请求停止，将在当前订单结束前停止")
                finalize_order_folder(folder, folder_name, failure_records)
                return 130
            emit(
                progress_callback,
                ProgressEvent(
                    order_index=order_index,
                    order_total=order_total,
                    order_no=folder_name,
                    item_index=item_index,
                    item_total=item_total,
                    message=f"正在下载 {item_index}/{item_total}",
                ),
            )
            print(f"  下载类别：{category}")
            browser_used = False
            direct_error_msg = ""
            try:
                try:
                    files = download_url(url, folder, download_index)
                except Exception as direct_error:
                    direct_error_msg = str(direct_error)
                    if args.browser_fallback == "off":
                        raise
                    is_google_drive = urlparse(url).netloc.lower().endswith("drive.google.com")
                    if is_google_drive and args.browser_fallback == "headless":
                        raise RuntimeError(
                            f"Google Drive 自动下载失败：{direct_error}。"
                            "为避免无头浏览器等待登录而阻塞，已跳过浏览器回退；"
                            "原始链接将写入 link.txt"
                        ) from direct_error
                    mode = "无头" if args.browser_fallback == "headless" else "可见"
                    browser_used = True
                    print(f"    直接下载失败，启动{mode}浏览器回退：{direct_error}")
                    print(f"    浏览器回退最长等待 {BROWSER_TIMEOUT} 秒：{url}")
                    files = download_with_browser(
                        url,
                        folder,
                        Path(args.browser_profile),
                        headless=args.browser_fallback == "headless",
                    )
                    print("    浏览器回退成功")
                for downloaded in files:
                    try:
                        extracted = safe_extract_archive(downloaded)
                        print(
                            f"  已下载：{downloaded.name}"
                            + (f"，已解压到 {extracted.name}" if extracted else "")
                        )
                    except Exception as extract_error:
                        failures += 1
                        failure_records.append(
                            DownloadFailure(
                                order_no=folder_name,
                                category=category,
                                url=url,
                                reason=str(extract_error),
                                stage="extract",
                                file_name=downloaded.name,
                            )
                        )
                        print(
                            f"  解压失败：订单={folder_name} 类别={category} 文件={downloaded.name}\n"
                            f"    原因：{extract_error}",
                            file=sys.stderr,
                        )
            except Exception as exc:
                failures += 1
                failure_records.append(
                    DownloadFailure(
                        order_no=folder_name,
                        category=category,
                        url=url,
                        reason=str(exc),
                        browser_used=browser_used,
                        direct_error=direct_error_msg if browser_used else "",
                    )
                )
                print(
                    f"  下载失败：订单={folder_name} 类别={category}\n"
                    f"    URL={url}\n"
                    f"    原因：{exc}\n"
                    f"    已启动浏览器回退：{'是' if browser_used else '否'}\n"
                    f"    将写入 link.txt：是",
                    file=sys.stderr,
                )
            download_index += 1

        if not args.skip_downloads:
            folder = finalize_order_folder(folder, folder_name, failure_records)
            if any(item.stage == "download" and item.url for item in failure_records):
                print(
                    f"  失败提醒：目录已改名为 {folder.name}，"
                    f"失败链接已写入 {LINK_FILE_NAME}，详细原因已写入 {FAILURE_LOG_NAME}"
                )
        emit(
            progress_callback,
            ProgressEvent(
                order_index=order_index,
                order_total=order_total,
                order_no=folder_name,
                item_index=item_total,
                item_total=item_total,
                message=f"完成 {folder.name}",
            ),
        )

    print(f"完成：输出目录 {output_root.resolve()}，失败项 {failures}")
    return 1 if failures else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="输入的 xlsx 文件")
    parser.add_argument("-o", "--output", type=Path, default=Path("output"), help="输出目录（默认：output）")
    parser.add_argument("--sheet", help="工作表名称（默认使用当前活动表）")
    parser.add_argument("--limit", type=int, help="只处理前 N 条数据，用于测试")
    parser.add_argument("--skip-downloads", action="store_true", help="只生成目录、文本和二维码，不下载素材")
    parser.add_argument(
        "--browser-fallback",
        choices=("off", "headless", "visible"),
        default="visible",
        help="直连失败后的浏览器回退方式（默认：visible）",
    )
    parser.add_argument(
        "--browser-profile",
        type=Path,
        default=Path(".xlsx2files-browser-profile"),
        help="浏览器会话目录；visible 模式登录一次后可供后续复用",
    )
    return parser


if __name__ == "__main__":
    try:
        raise SystemExit(process(build_parser().parse_args()))
    except KeyboardInterrupt:
        print("用户中止", file=sys.stderr)
        raise SystemExit(130)
    except Exception as error:
        print(f"错误：{error}", file=sys.stderr)
        raise SystemExit(1)
